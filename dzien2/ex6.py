import openpyxl

wb = openpyxl.load_workbook('video2.xlsx')

ws = wb['vgsales']

ws['P1'] = 'Average Sales'
# ws['P2'] = '= AVERAGE(K2:K16199)'
ws['P2'] = '= AVERAGE(L2:L16199)'  # =ŚREDNIA(L2:L16199)

wb.save('video2.xlsx')
wb.close()

ws['Q1'] = 'Number of populated cells'
ws['Q2'] = "=COUNTA(E2:E16220)"

wb.save('video2.xlsx')
wb.close()

ws['R1'] = "Number of rows with Sports Genre"
ws['R2'] = '=COUNTIF(E2:E16620, "sports")'

wb.save('video2.xlsx')
wb.close()

ws['S1'] = 'Total sports Sales'
# ws['S2'] = '=SUMIF(E2:E16220, "Sports", K2:K16220)'
ws['S2'] = '=SUMIF(E2:E16220, "Sports", L2:L16220)'

wb.save('video2.xlsx')
wb.close()

ws['T1'] = "Rounded Sum of Sports Sales"
ws['T2'] = '=CEILING(S2, 25)'  # do najblizszej 25, 1308.9 -> 1325

wb.save('video2.xlsx')
wb.close()

ws['U1'] = "Rounded Sum of Sports Sales"
ws['U2'] = '=CEILING(S2, 1)'  # do najblizszej całkowitej, 1308.9 -> 1309

wb.save('video2.xlsx')
wb.close()
