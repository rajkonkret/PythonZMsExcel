import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('video2.xlsx')
ws = wb.active

# print(ws.title)
# # zmiana nazwy arkusza
# ws.title = "Video Games Sales Data"
# wb.save("video2.xlsx")
# wb.close()

# # lista arkuszy
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']
# sheet_names = wb.sheetnames
# print(sheet_names)
# # ustawienie aktywnego arkusza po indeksie
# sheet = wb[sheet_names[1]]  # indeks 1 czyli -> Total Sales by Genre
# wb.active = wb.index(sheet)
#
# ws = wb.active
# print(ws.title)  # Total Sales by Genre
# ctrl / - komentowanie

# formatowanie tekstu w pierwszym wierszu
# ws = wb['Video Games Sales Data']
# ws['A1'].font = Font(bold=True, size=12)
#
# for cell in ws[1:1]:
#     cell.font = Font(bold=True, size=12)
#
# wb.save('video2.xlsx')
# wb.close()

# dodanie arkusza do pliku
print(wb.sheetnames)
# wb.create_sheet("Empty Sheet")
# print(wb.sheetnames)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year', 'Empty Sheet']
#
# wb.save('video2.xlsx')
# wb.close()

# usunięcie arkusza z pliku
# wb.remove(wb['Empty Sheet'])
# print(wb.sheetnames)
# wb.save('video2.xlsx')
# wb.close()

# kopiowanie arkusza
wb.copy_worksheet(wb['Video Games Sales Data'])
wb.save('video2_kopia.xlsx')
wb.close()
