import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('video2.xlsx')
ws = wb.active

ws = wb['Video Games Sales Data']

# FF0000 -> RGB -> R=FF G=0 B=0
ws['A1'].font = Font(color='FF0000', bold=True, size=12)  # red
ws['A2'].font = Font(color='0000FF')  # blue

ws['A1'].fill = PatternFill('lightVertical', start_color='38e3ff')
# FILL_NONE = 'none'
# FILL_SOLID = 'solid'
# FILL_PATTERN_DARKDOWN = 'darkDown'
# FILL_PATTERN_DARKGRAY = 'darkGray'
# FILL_PATTERN_DARKGRID = 'darkGrid'
# FILL_PATTERN_DARKHORIZONTAL = 'darkHorizontal'
# FILL_PATTERN_DARKTRELLIS = 'darkTrellis'
# FILL_PATTERN_DARKUP = 'darkUp'
# FILL_PATTERN_DARKVERTICAL = 'darkVertical'
# FILL_PATTERN_GRAY0625 = 'gray0625'
# FILL_PATTERN_GRAY125 = 'gray125'
# FILL_PATTERN_LIGHTDOWN = 'lightDown'
# FILL_PATTERN_LIGHTGRAY = 'lightGray'
# FILL_PATTERN_LIGHTGRID = 'lightGrid'
# FILL_PATTERN_LIGHTHORIZONTAL = 'lightHorizontal'
# FILL_PATTERN_LIGHTTRELLIS = 'lightTrellis'
# FILL_PATTERN_LIGHTUP = 'lightUp'
# FILL_PATTERN_LIGHTVERTICAL = 'lightVertical'  ----
# FILL_PATTERN_MEDIUMGRAY = 'mediumGray'

# ramka
my_border = Side(border_style='thin', color='000000')  # czarny
ws['A1'].border = Border(
    top=my_border, left=my_border, right=my_border, bottom=my_border
)

# oznaczenie wyfiltrowanych danych
fill = PatternFill(
    start_color='90EE90',
    end_color='90EE90',
    fill_type="solid"
)

# {">": "greaterThan", ">=": "greaterThanOrEqual", "<": "lessThan", "<=": "lessThanOrEqual",
#               "=": "equal", "==": "equal", "!=": "notEqual"}
ws.conditional_formatting.add(
    'G2:L16599',
    CellIsRule(operator="lessThan",
               formula=[5],
               fill=fill,
               font=Font(color='FF00FF'))
)

fill_gt = PatternFill(
    start_color='70EE90',
    end_color='70EE90',
    fill_type="solid"
)

# ["solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal", "darkTrellis", "darkUp", "darkVertical", "gray0625",
#  "gray125", "lightDown", "lightGray", "lightGrid", "lightHorizontal", "lightTrellis", "lightUp", "lightVertical",
#  "mediumGray", "none"]
ws.conditional_formatting.add(
    'G2:L16599',
    CellIsRule(operator="greaterThan",
               formula=[8],
               fill=fill_gt,
               font=Font(color='FF0000'))
)

wb.save('video2.xlsx')
wb.close()
