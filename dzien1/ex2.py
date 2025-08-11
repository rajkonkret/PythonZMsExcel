import openpyxl
from pprint import pprint

# katalog w głownym, wychodzimy z dzien1 i wchodzimy do data
# .. - wyjdz wyzej
# wb = openpyxl.load_workbook("../data/videogamesales.xlsx")
wb = openpyxl.load_workbook("../data/videogamesales.xlsx")
ws = wb.active  # aktywny arkusz

print(wb)
print(ws)  # <Worksheet "vgsales"> - nazwa arkusza: vgsales

# gdy znamy arkusze w pliku
# możemy ustawic konkretny do odczytu
ws = wb['vgsales']

print("Total number of row:", ws.max_row)  # Total number of row: 16328
print("Total number of columns:", ws.max_column)  # Total number of columns: 10

print("Value in cell A1 is:", ws['A1'].value)

# wczytanie wartości z pierwszego wiersza dla wszystkich kolumn
# umiescimy te dane w liscie []
# +1 aby wziął ostatnią kolumnę
values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
print(values)
# dostaliśmy nazwy kolumn
# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales']
# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales', 'Other_Sales']

# Dane z kolumny 2, wiersze od 2 do 11 (12 nie brany pod uwage - tak działą python)
data = [ws.cell(row=i, column=2).value for i in range(2, 12)]
print(data)
# ['Wii Sports',
#  'Super Mario Bros.',
#  'Mario Kart Wii',
#  'Wii Sports Resort',
#  'Pokemon Red/Pokemon Blue',
#  'Tetris', 'New Super Mario Bros.',
#  'Wii Play',
#  'New Super Mario Bros. Wii',
#  'Duck Hunt']

my_list = list()  # pusta lista

for value in ws.iter_rows(
        min_row=1, max_row=11, min_col=1, max_col=6,
        values_only=True):
    my_list.append(value)

# print(my_list)
pprint(my_list)
# [('Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher'),
#  (1, 'Wii Sports', 'Wii', 2006, 'Sports', 'Nintendo'),
#  (2, 'Super Mario Bros.', 'NES', 1985, 'Platform', 'Nintendo'),
#  (3, 'Mario Kart Wii', 'Wii', 2008, 'Racing', 'Nintendo'),
#  (4, 'Wii Sports Resort', 'Wii', 2009, 'Sports', 'Nintendo'),
#  (5, 'Pokemon Red/Pokemon Blue', 'GB', 1996, 'Role-Playing', 'Nintendo'),
#  (6, 'Tetris', 'GB', 1989, 'Puzzle', 'Nintendo'),
#  (7, 'New Super Mario Bros.', 'DS', 2006, 'Platform', 'Nintendo'),
#  (8, 'Wii Play', 'Wii', 2006, 'Misc', 'Nintendo'),
#  (9, 'New Super Mario Bros. Wii', 'Wii', 2009, 'Platform', 'Nintendo'),
#  (10, 'Duck Hunt', 'NES', 1984, 'Shooter', 'Nintendo')]