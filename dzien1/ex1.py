from openpyxl import Workbook, load_workbook

# ewentualna instalacja biblioteki w consoli
# pip install openpyxl

wb = Workbook()  # tworzymy excel
ws = wb.active  # ustawiamy aktywny arkusz

# komórka w arkuszu A1
ws['A1'] = 42

ws.append([1, 2, 3])  # lista int

# zapisanie danych do pliku excel
wb.save('sample.xlsx')
wb.close()  # zamknięcie pliku

print("File name: sample.xlsx")
print("Location in current directory")
# File name: sample.xlsx
# Location in current directory

# wczytujemy plik excel do pamięci
workbook = load_workbook("sample.xlsx")
sheet = workbook.active  # ustawiamy aktywny arkusz
print(sheet)  # wyswietlamu nazwę tego arkusza, <Worksheet "Sheet">
print(sheet['A1'].value)  # wypisanie wartości z komórki A1, tu: 42

# wypisanie wierszy excela
for row in sheet.iter_rows(min_row=1, max_row=3):  # wiersze
    for cell in row:  # wszystkie kolumny
        print(cell.value)
# NaN -> None
